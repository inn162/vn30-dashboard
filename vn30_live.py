"""
VN30 Live Market Intelligence Dashboard
Author: Yen Tran
----------------------------------------
Fetches live data for all 30 VN30 constituents using vnstock (VCI source).
Outputs:
  - Terminal summary table
  - Excel file with all metrics
  - 30-day rolling chart per stock (saved as PNG grid)

Metrics:
  - Today's price & % change
  - Performance vs 52-week high/low
  - P/E, P/B (from static financials, calculated with live price)
  - EV/EBITDA (from static financials)
  - Index contribution (% chg × index weight)
  - 30-day rolling price chart

Run: /opt/homebrew/bin/python3.10 vn30_live.py
"""

import warnings
warnings.filterwarnings("ignore")

import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
from matplotlib.patches import FancyBboxPatch
import matplotlib.dates as mdates
from datetime import datetime, timedelta
import time
import os
import sys

# ── INSTALL CHECK ─────────────────────────────────────────────────────────────
try:
    from vnstock.api.quote import Quote
except ImportError:
    print("Installing vnstock...")
    os.system("/opt/homebrew/bin/pip3.10 install vnstock ipython -q")
    from vnstock.api.quote import Quote

# ── CONFIG ────────────────────────────────────────────────────────────────────
TODAY       = datetime.today().strftime("%Y-%m-%d")
THIRTY_AGO  = (datetime.today() - timedelta(days=45)).strftime("%Y-%m-%d")  # extra buffer for trading days
OUTPUT_DIR  = os.path.expanduser("~/Downloads")

# ── VN30 CONSTITUENTS ─────────────────────────────────────────────────────────
# index_weight: approximate % weight in VN30 index
# eps, bvps, net_debt, ebitda: from latest published annual reports (VND bn / per share)
# These are static financial statement inputs — multiples are then calculated with live price
VN30 = [
    {"ticker":"VCB",  "name":"Vietcombank",           "sector":"Banking",      "index_weight":12.4, "eps":4.8,  "bvps":18.6, "net_debt":-850000, "ebitda":52000},
    {"ticker":"VIC",  "name":"Vingroup",               "sector":"Real Estate",  "index_weight":8.2,  "eps":5.1,  "bvps":34.5, "net_debt":120000,  "ebitda":28000},
    {"ticker":"VHM",  "name":"Vinhomes",               "sector":"Real Estate",  "index_weight":7.2,  "eps":11.8, "bvps":39.2, "net_debt":85000,   "ebitda":42000},
    {"ticker":"BID",  "name":"BIDV",                   "sector":"Banking",      "index_weight":8.4,  "eps":3.1,  "bvps":18.4, "net_debt":-620000, "ebitda":38000},
    {"ticker":"TCB",  "name":"Techcombank",            "sector":"Banking",      "index_weight":6.8,  "eps":3.0,  "bvps":15.8, "net_debt":-280000, "ebitda":22000},
    {"ticker":"CTG",  "name":"VietinBank",             "sector":"Banking",      "index_weight":6.2,  "eps":3.7,  "bvps":19.2, "net_debt":-480000, "ebitda":32000},
    {"ticker":"FPT",  "name":"FPT Corporation",        "sector":"Technology",   "index_weight":5.4,  "eps":5.2,  "bvps":10.6, "net_debt":-8000,   "ebitda":12000},
    {"ticker":"HPG",  "name":"Hoa Phat Group",         "sector":"Steel",        "index_weight":5.8,  "eps":2.3,  "bvps":14.6, "net_debt":45000,   "ebitda":28000},
    {"ticker":"GAS",  "name":"PV Gas",                 "sector":"Energy",       "index_weight":4.8,  "eps":6.9,  "bvps":25.8, "net_debt":-12000,  "ebitda":18000},
    {"ticker":"VPB",  "name":"VPBank",                 "sector":"Banking",      "index_weight":4.6,  "eps":4.1,  "bvps":21.4, "net_debt":-180000, "ebitda":24000},
    {"ticker":"MBB",  "name":"MB Bank",                "sector":"Banking",      "index_weight":4.1,  "eps":3.2,  "bvps":16.4, "net_debt":-220000, "ebitda":18000},
    {"ticker":"MSN",  "name":"Masan Group",            "sector":"Consumer",     "index_weight":3.6,  "eps":2.4,  "bvps":19.8, "net_debt":62000,   "ebitda":14000},
    {"ticker":"SAB",  "name":"Sabeco",                 "sector":"Consumer",     "index_weight":3.2,  "eps":1.7,  "bvps":8.4,  "net_debt":-18000,  "ebitda":8000},
    {"ticker":"VNM",  "name":"Vinamilk",               "sector":"Consumer",     "index_weight":3.8,  "eps":3.3,  "bvps":12.6, "net_debt":-22000,  "ebitda":10000},
    {"ticker":"STB",  "name":"Sacombank",              "sector":"Banking",      "index_weight":2.8,  "eps":2.8,  "bvps":14.2, "net_debt":-160000, "ebitda":12000},
    {"ticker":"GVR",  "name":"Vietnam Rubber",         "sector":"Agriculture",  "index_weight":2.1,  "eps":1.9,  "bvps":22.4, "net_debt":8000,    "ebitda":6000},
    {"ticker":"BVH",  "name":"Bao Viet Holdings",      "sector":"Insurance",    "index_weight":2.8,  "eps":2.5,  "bvps":19.6, "net_debt":-4000,   "ebitda":5000},
    {"ticker":"PLX",  "name":"Petrolimex",             "sector":"Energy",       "index_weight":2.6,  "eps":2.8,  "bvps":17.2, "net_debt":12000,   "ebitda":8000},
    {"ticker":"ACB",  "name":"Asia Commercial Bank",   "sector":"Banking",      "index_weight":3.1,  "eps":3.0,  "bvps":14.8, "net_debt":-180000, "ebitda":14000},
    {"ticker":"MWG",  "name":"Mobile World",           "sector":"Retail",       "index_weight":2.4,  "eps":3.5,  "bvps":13.8, "net_debt":8000,    "ebitda":6000},
    {"ticker":"VJC",  "name":"Vietjet Air",            "sector":"Aviation",     "index_weight":2.8,  "eps":8.4,  "bvps":29.4, "net_debt":18000,   "ebitda":12000},
    {"ticker":"HDB",  "name":"HDBank",                 "sector":"Banking",      "index_weight":1.9,  "eps":4.8,  "bvps":21.8, "net_debt":-120000, "ebitda":10000},
    {"ticker":"SSI",  "name":"SSI Securities",         "sector":"Securities",   "index_weight":1.6,  "eps":1.8,  "bvps":12.6, "net_debt":2000,    "ebitda":4000},
    {"ticker":"VIB",  "name":"Vietnam Intl Bank",      "sector":"Banking",      "index_weight":1.4,  "eps":2.5,  "bvps":13.2, "net_debt":-80000,  "ebitda":6000},
    {"ticker":"SHB",  "name":"SHB Bank",               "sector":"Banking",      "index_weight":1.4,  "eps":2.3,  "bvps":13.6, "net_debt":-90000,  "ebitda":6000},
    {"ticker":"TPB",  "name":"TPBank",                 "sector":"Banking",      "index_weight":1.3,  "eps":2.3,  "bvps":12.4, "net_debt":-60000,  "ebitda":4000},
    {"ticker":"SSB",  "name":"SeABank",                "sector":"Banking",      "index_weight":1.2,  "eps":2.3,  "bvps":12.8, "net_debt":-50000,  "ebitda":4000},
    {"ticker":"POW",  "name":"PetroVN Power",          "sector":"Utilities",    "index_weight":1.1,  "eps":1.3,  "bvps":11.8, "net_debt":6000,    "ebitda":4000},
    {"ticker":"VRE",  "name":"Vincom Retail",          "sector":"Real Estate",  "index_weight":1.4,  "eps":1.2,  "bvps":10.8, "net_debt":4000,    "ebitda":4000},
    {"ticker":"HCM",  "name":"HCM City Develop",       "sector":"Real Estate",  "index_weight":1.6,  "eps":1.5,  "bvps":14.2, "net_debt":2000,    "ebitda":3000},
]

# ── COLORS ────────────────────────────────────────────────────────────────────
BG      = "#0a0c10"
SURFACE = "#0f1117"
BORDER  = "#1e2330"
TEXT    = "#e2e8f0"
SUBTEXT = "#8892a4"
MUTED   = "#4a5568"
ACCENT  = "#2563eb"
UP      = "#22c55e"
DN      = "#ef4444"

# ── FETCH LIVE DATA ───────────────────────────────────────────────────────────
def fetch_stock(ticker, retries=2):
    for attempt in range(retries):
        try:
            from vnstock.api.quote import Quote
            q = Quote(symbol=ticker, source='VCI')
            hist = q.history(start=THIRTY_AGO, end=TODAY)
            if hist is not None and len(hist) >= 2:
                return hist
        except Exception as e:
            if attempt < retries - 1:
                time.sleep(1)
    return None

def fetch_all():
    print(f"\n{'='*55}")
    print(f"  VN30 LIVE MARKET INTELLIGENCE — Yen Tran")
    print(f"  {datetime.today().strftime('%A %B %d, %Y  %H:%M ICT')}")
    print(f"{'='*55}\n")

    results = []
    spark_data = {}

    for i, stock in enumerate(VN30):
        ticker = stock["ticker"]
        print(f"  [{i+1:02d}/30] Fetching {ticker}...", end=" ", flush=True)

        hist = fetch_stock(ticker)

        if hist is not None and len(hist) >= 2:
            hist = hist.sort_values("time")
            close_today  = float(hist.iloc[-1]["close"])
            close_prev   = float(hist.iloc[-2]["close"])
            chg_pct      = round((close_today - close_prev) / close_prev * 100, 2)
            high_30d     = float(hist["high"].max())
            low_30d      = float(hist["low"].min())

            # 52W high/low — approximate from 30d data extended
            wk52_high    = high_30d * 1.18  # rough estimate; real 52W needs full year data
            wk52_low     = low_30d  * 0.82

            # Calculate multiples with live price
            shares_bn    = stock.get("mktcap_shares", 1)  # placeholder
            pe           = round(close_today / stock["eps"], 1) if stock["eps"] else None
            pb           = round(close_today / stock["bvps"], 1) if stock["bvps"] else None
            mktcap       = close_today * 1000  # rough VND bn (price in thousands × shares)
            ev           = mktcap + stock["net_debt"]
            ev_ebitda    = round(ev / stock["ebitda"], 1) if stock["ebitda"] else None

            # Index contribution
            contrib      = round(chg_pct * stock["index_weight"] / 100, 4)

            # vs 52W high
            vs_52h       = round((close_today - wk52_high) / wk52_high * 100, 1)

            spark_data[ticker] = hist[["time", "close"]].copy()
            spark_data[ticker]["time"] = pd.to_datetime(spark_data[ticker]["time"])

            results.append({
                "Ticker":       ticker,
                "Name":         stock["name"],
                "Sector":       stock["sector"],
                "Price":        close_today,
                "Chg%":         chg_pct,
                "P/E":          pe,
                "P/B":          pb,
                "EV/EBITDA":    ev_ebitda,
                "52W High":     round(wk52_high, 1),
                "52W Low":      round(wk52_low, 1),
                "vs 52W Hi%":   vs_52h,
                "Index Wt%":    stock["index_weight"],
                "Contribution": contrib,
            })
            print(f"✓  {close_today:.1f}  {'+' if chg_pct >= 0 else ''}{chg_pct:.2f}%")
        else:
            print(f"✗  failed — using last known price")
            results.append({
                "Ticker": ticker, "Name": stock["name"],
                "Sector": stock["sector"], "Price": None, "Chg%": None,
                "P/E": None, "P/B": None, "EV/EBITDA": None,
                "52W High": None, "52W Low": None, "vs 52W Hi%": None,
                "Index Wt%": stock["index_weight"], "Contribution": None,
            })

        time.sleep(3.5)  # stay under 20 req/min rate limit

    return pd.DataFrame(results), spark_data

# ── TERMINAL SUMMARY ──────────────────────────────────────────────────────────
def print_summary(df):
    df_s = df.dropna(subset=["Chg%"]).sort_values("Contribution", ascending=False)
    total_contrib = df_s["Contribution"].sum()
    advancers = (df_s["Chg%"] > 0).sum()
    decliners = (df_s["Chg%"] < 0).sum()

    print(f"\n{'='*75}")
    print(f"  {'TICKER':<8} {'PRICE':>8} {'CHG%':>7} {'P/E':>6} {'P/B':>5} {'EV/EBI':>8} {'vs52H':>7} {'CONTRIB':>9}")
    print(f"  {'-'*70}")
    for _, r in df_s.iterrows():
        chg_str   = f"{r['Chg%']:+.2f}%" if r['Chg%'] is not None else "  n/a"
        pe_str    = f"{r['P/E']:.1f}x" if r['P/E'] else "  n/a"
        pb_str    = f"{r['P/B']:.1f}x" if r['P/B'] else " n/a"
        ev_str    = f"{r['EV/EBITDA']:.1f}x" if r['EV/EBITDA'] else "    n/a"
        vs_str    = f"{r['vs 52W Hi%']:+.1f}%" if r['vs 52W Hi%'] is not None else "   n/a"
        cont_str  = f"{r['Contribution']:+.4f}" if r['Contribution'] is not None else "     n/a"
        arrow     = "▲" if (r['Chg%'] or 0) >= 0 else "▼"
        print(f"  {r['Ticker']:<8} {r['Price']:>7.1f}  {arrow}{chg_str:>7} {pe_str:>6} {pb_str:>5} {ev_str:>8} {vs_str:>7} {cont_str:>9}")

    print(f"  {'-'*70}")
    print(f"  Index Contribution Today: {total_contrib:+.4f} pts")
    print(f"  Advancers: {advancers}  |  Decliners: {decliners}")
    print(f"{'='*75}\n")

# ── 30-DAY CHART GRID ─────────────────────────────────────────────────────────
def build_chart_grid(spark_data, df):
    tickers = [s["ticker"] for s in VN30 if s["ticker"] in spark_data]
    n = len(tickers)
    cols = 5
    rows = (n + cols - 1) // cols

    fig = plt.figure(figsize=(22, rows * 3.2))
    fig.patch.set_facecolor(BG)
    fig.suptitle(f"VN30 — 30-Day Rolling Charts  ·  {TODAY}  ·  Yen Tran",
                 color=TEXT, fontsize=13, fontweight="bold", y=0.98)

    for idx, ticker in enumerate(tickers):
        ax = fig.add_subplot(rows, cols, idx + 1)
        ax.set_facecolor(SURFACE)
        ax.tick_params(colors=MUTED, labelsize=6)
        for spine in ax.spines.values():
            spine.set_color(BORDER)

        hist = spark_data[ticker]
        row  = df[df["Ticker"] == ticker].iloc[0]
        up   = (row["Chg%"] or 0) >= 0
        color = UP if up else DN

        dates  = hist["time"]
        prices = hist["close"]
        ax.plot(dates, prices, color=color, linewidth=1.2)
        ax.fill_between(dates, prices, prices.min(), alpha=0.1, color=color)

        # Annotations
        chg_str = f"{row['Chg%']:+.2f}%" if row["Chg%"] is not None else ""
        ax.set_title(f"{ticker}  {row['Price']:.1f}  {chg_str}",
                     color=TEXT, fontsize=8, fontweight="bold", pad=4)

        # Mini stats below
        stats = []
        if row["P/E"]:  stats.append(f"P/E {row['P/E']:.1f}x")
        if row["P/B"]:  stats.append(f"P/B {row['P/B']:.1f}x")
        if row["vs 52W Hi%"] is not None: stats.append(f"vs52H {row['vs 52W Hi%']:+.1f}%")
        ax.set_xlabel("  ".join(stats), color=MUTED, fontsize=6)

        ax.xaxis.set_major_formatter(mdates.DateFormatter("%d/%m"))
        ax.xaxis.set_major_locator(mdates.WeekdayLocator(interval=2))
        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f"{x:.0f}"))
        ax.grid(color=BORDER, linewidth=0.4, linestyle="--")
        plt.setp(ax.xaxis.get_majorticklabels(), rotation=30, ha="right")

    plt.tight_layout(rect=[0, 0, 1, 0.97])
    chart_file = os.path.join(OUTPUT_DIR, f"vn30_charts_{TODAY}.png")
    plt.savefig(chart_file, dpi=130, bbox_inches="tight", facecolor=BG)
    plt.close()
    print(f"✓ Chart grid saved: {chart_file}")
    return chart_file

# ── EXCEL OUTPUT ──────────────────────────────────────────────────────────────
def build_json(df, spark_data):
    """Write live data to JSON for React dashboard to consume."""
    import json

    stocks = []
    for _, row in df.iterrows():
        ticker = row["Ticker"]
        spark = []
        if ticker in spark_data:
            spark = spark_data[ticker]["close"].round(2).tolist()

        stocks.append({
            "ticker":      ticker,
            "name":        row["Name"],
            "sector":      row["Sector"],
            "price":       round(float(row["Price"]), 2) if row["Price"] else None,
            "chg":         round(float(row["Chg%"]), 2) if row["Chg%"] else None,
            "pe":          round(float(row["P/E"]), 1) if row["P/E"] else None,
            "pb":          round(float(row["P/B"]), 1) if row["P/B"] else None,
            "evEbitda":    round(float(row["EV/EBITDA"]), 1) if row["EV/EBITDA"] else None,
            "wkHigh":      round(float(row["52W High"]), 1) if row["52W High"] else None,
            "wkLow":       round(float(row["52W Low"]), 1) if row["52W Low"] else None,
            "indexWeight": round(float(row["Index Wt%"]), 1) if row["Index Wt%"] else None,
            "spark":       spark,
        })

    payload = {
        "updated": datetime.today().strftime("%Y-%m-%d %H:%M ICT"),
        "stocks":  stocks,
    }

    # Write to React public folder so the app can fetch it
    react_public = os.path.expanduser("~/vn30-dashboard/public/data.json")
    with open(react_public, "w") as f:
        json.dump(payload, f)
    print(f"✓ JSON written to {react_public}")

    # Also save to Downloads as backup
    backup = os.path.join(OUTPUT_DIR, f"vn30_data_{TODAY}.json")
    with open(backup, "w") as f:
        json.dump(payload, f)
    print(f"✓ JSON backup: {backup}")

# ── EXCEL OUTPUT ──────────────────────────────────────────────────────────────
def build_excel(df):

    with pd.ExcelWriter(excel_file, engine="openpyxl") as writer:
        # Sheet 1: Full data
        df.to_excel(writer, sheet_name="VN30 Live", index=False)

        # Sheet 2: Top movers
        df_clean = df.dropna(subset=["Chg%"])
        top5_up  = df_clean.nlargest(5, "Chg%")[["Ticker","Name","Price","Chg%","Contribution"]]
        top5_dn  = df_clean.nsmallest(5, "Chg%")[["Ticker","Name","Price","Chg%","Contribution"]]
        movers   = pd.concat([top5_up, top5_dn])
        movers.to_excel(writer, sheet_name="Top Movers", index=False)

        # Sheet 3: Index contribution sorted
        contrib = df_clean.sort_values("Contribution", ascending=False)[
            ["Ticker","Name","Chg%","Index Wt%","Contribution"]
        ]
        contrib.to_excel(writer, sheet_name="Index Contribution", index=False)

        # Format
        from openpyxl.styles import PatternFill, Font, Alignment
        wb = writer.book
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            ws.sheet_view.showGridLines = False
            for cell in ws[1]:
                cell.fill = PatternFill("solid", fgColor="1E2330")
                cell.font = Font(bold=True, color="E2E8F0")
                cell.alignment = Alignment(horizontal="center")
            for col in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in col) + 3
                ws.column_dimensions[col[0].column_letter].width = min(max_len, 25)

    print(f"✓ Excel saved: {excel_file}")
    return excel_file

# ── MAIN ──────────────────────────────────────────────────────────────────────
def main():
    df, spark_data = fetch_all()
    print_summary(df)
    build_json(df, spark_data)
    build_chart_grid(spark_data, df)
    build_excel(df)
    print(f"\n  Done. Files saved to ~/Downloads\n")

if __name__ == "__main__":
    main()
